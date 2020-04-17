

def append_data_last_sheet(path, min_row_headers, min_row_data, dest_filename):

    import openpyxl as xl
    wb1 = xl.load_workbook(path)

    wb1.create_sheet('CombineHere')

    sheets = wb1.sheetnames
    last_sheet = wb1[sheets[-1]]

    list_len=len(sheets[:-1])
    print("Combine function")
    for j in range(list_len):

        for row in wb1[sheets[j]].iter_rows(min_row=min_row_data if j!=0 else min_row_headers, max_row=wb1[sheets[j]].max_row, min_col=None, max_col=None, values_only=False):

            last_sheet.append(cell.value for cell in row)

    wb1.save(dest_filename)


def add_tla_first_col(path, min_row_data, dest_filename):
    import openpyxl as xl
    wb1 = xl.load_workbook(path)

    sheets = wb1.sheetnames

    list_len = len(sheets)
    print("Add function")
    for j in range(list_len):
        tla_whole = wb1[sheets[j]].cell(row=5, column=1).value
        tla = tla_whole.split()

        wb1[sheets[j]].insert_cols(1,2)

        wb1[sheets[j]].cell(row=9, column=1).value = "TLA"
        wb1[sheets[j]].cell(row=9, column=2).value = "#ofEntries"

        for row in wb1[sheets[j]].iter_rows(min_row=min_row_data, max_row=wb1[sheets[j]].max_row, min_col=1, max_col=1,values_only=False):
            for cell in row:
                cell.value = tla[1]


        for row in wb1[sheets[j]].iter_rows(min_row=min_row_data, max_row=wb1[sheets[j]].max_row, min_col=2, max_col=2,values_only=False):
            for cell in row:
                cell.value = wb1[sheets[j]].max_row-min_row_data+1

        print("Done with [",j,"]")

    wb1.save(dest_filename)


def append_data_new_wb_ro(path, min_row_headers, min_row_data, dest_filename):
    import openpyxl as xl
    wb1 = xl.load_workbook(path,read_only=True)
    wb2 = xl.Workbook(write_only=True)
    wb2_ws1=wb2.create_sheet()
    sheets = wb1.sheetnames

    print("Combine function read only")
    for j in range(len(sheets)):

        for row in wb1[sheets[j]].iter_rows(min_row=min_row_data if j != 0 else min_row_headers,max_row=wb1[sheets[j]].max_row, min_col=None, max_col=None,values_only=False):
            wb2_ws1.append(cell.value for cell in row)
        print("Done with sheet[",j,"]")

    wb2.save(dest_filename)

dest_filename = '/Users/vchaubal/Documents/Exploded BOMs for Systems/Combined_BOM_4_16_3.xlsx'
path = '/Users/vchaubal/Documents/Exploded BOMs for Systems/Combined_BOM_4_16_2.xlsx'
min_row_headers=9
max_row_headers=1
min_row_data = 10
#add_tla_first_col(path,min_row_data,dest_filename)
append_data_new_wb_ro(path,min_row_headers,min_row_data,dest_filename)